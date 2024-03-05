import datetime
import os
import subprocess
from urllib.parse import quote_plus

import pandas as pd
import pyodbc
import openpyxl
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import create_engine
import sys

script_name = sys.argv[0]
Vendor_Number = sys.argv[1]
Year = sys.argv[2]

#Vendor_Number = input("Enter Vendor Number: ")
#Year = input("Enter Year: ")

params = {
    'Vendor_no': Vendor_Number,
    'Year': Year,
    'Member_Number': None,
}

#Define your first SQL Server Connection Parameters
#Credentials removed for privacy

#Create a connection string for the first SQL Server
connection_string1 = f"mssql+pyodbc://{username1}:{password1}@{server1}/{database1}?driver=SQL+Server"

#Define your second SQL Server Connection Parameters
#Credentials removed for privacy

#Create a connection string for the second SQL Server
connection_string2 = f"mssql+pyodbc://@{server2}/{database2}?driver=SQL+Server"

#Define your third SQL Server Connection string for connection3 (Access database)
#Credentials removed for privacy

connection_string3 = f"mssql+pyodbc://{UID}:{PWD}@{DSN}"
print (f"Connection Successful, begin gathering Vendor Details for V# {Vendor_Number}")
try:
    # Create a SQLAlchemy engine for the first database connection
    engine1 = create_engine(connection_string1)

    #Create a SQLAlchemy engine for the second database connection
    engine2 = create_engine(connection_string2)

    #Establish a connection to the third Access database
    engine3 = create_engine(connection_string3)

    # Execute SQL query to pull APDunsNo for Vendor from tblVendorPR table
    apdunsno_query = "Select DunsNo from tblVendorPR where VendorNo = ?"

    apdunsno = pd.read_sql(apdunsno_query, engine1, params=(Vendor_Number,))

    if not apdunsno.empty:
        Member_Number = apdunsno.iloc[0]['DunsNo']
        params['Member_Number'] = Member_Number
    else:
        print("Member_number not found in Vendor_Pr for the given Vendor Number.")

    summary_query = f""" 
    select ProgramType as Programs, Percentage as '%', OSOProgramType as 'Program Type',
     CONCAT('COL# ', ContractID,' start ',CONVERT(DATE, ProgramStartDate),
        CASE
            WHEN CloseDate IS NOT NULL THEN CONCAT(' to ', CONVERT(DATE, CloseDate))
            WHEN CloseDate <= ProgramStartDate THEN CONCAT(' to ', CONVERT(DATE, CloseDate))
            ELSE ', Ongoing'
        END) AS 'Status'
    from (
    select VendorNo, ProgramType, ContractID, OSOProgramType, Percentage, Status, ProgramStartDate, CloseDate,
    ROW_NUMBER() OVER (PARTITION BY VENDORNO, OSOProgramType, YEAR(ProgramStartDate) ORDER BY Status ASC, PROGRAMSTARTDATE DESC) AS CONTRACTRANK
    from tblContracts  
    where OSOProgramType is not null and status in ('Active','Closed') and ((Year(ProgramStartDate) <= ? and (Year(CloseDate) >= ? or CloseDate is null)))
    ) AS RankedContracts
    where VendorNo = ?
    AND ProgramType = 'OSO' and CONTRACTRANK = 1
    order by VendorNo, ProgramStartDate
    """
    print("Gathering Vendor Details")
    details_query = f"""
SELECT 
        Store_Count, Member_Number, Addressee, Item_Number, Short_Description, Unit_Cost, Eligible_Qty,
        Eligible_Sales, DunsNo, Vendor_Number, Name, Year, PeriodAdded, Type, Vendor_Funding, Credit_Date 
    FROM (
        SELECT 
            A.Store_Count_Number AS Store_Count, A.Member_Number, D.ADDRESSEE AS Addressee,
            A.Item_Number, C.SHORT_DESCRIPTION AS Short_Description, 
            A.Unit_Cost as Unit_Cost, 
            ISNULL(A.Eligible_Qty, '') AS Eligible_Qty,
            A.Eligible_Sales AS Eligible_Sales,
            A.APDunsno as DunsNo, A.Vendor_Number AS Vendor_Number,
            A.Name, A.Year, A.PeriodAdded,
            CASE 
                WHEN E.OSOProgramType = 'NNM' THEN A.Project_Type
                WHEN E.OSOProgramType = 'GUP' THEN A.Type
                WHEN E.OSOProgramType = 'ORIG' THEN A.Type
                WHEN E.OSOProgramType IN ('GUP','ORIG') THEN A.Type
                ELSE A.Project_Type 
            END AS Type, 
            CASE 
                WHEN E.OSOProgramType = 'NNM' and A.Project_Type in ('M','NW') and CONCAT(A.Year, A.PeriodAdded) between EffPeriod and EffClosePeriod THEN A.Eligible_Sales * (E.Percentage / 100)
                WHEN E.OSOProgramType = 'GUP' AND A.Type IN ('G') and CONCAT(A.Year, A.PeriodAdded) between EffPeriod and EffClosePeriod THEN A.Eligible_Sales * (E.Percentage / 100)
                WHEN E.OSOProgramType = 'ORIG' AND A.Type IN ('N', 'O') and CONCAT(A.Year, A.PeriodAdded) between EffPeriod and EffClosePeriod THEN A.Eligible_Sales * (E.Percentage / 100)
            END AS Vendor_Funding,
            CONVERT(VARCHAR(10), A.Credit_Date, 101) AS Credit_Date 
        FROM 
            tblOSOBase A 
        LEFT JOIN 
           tblItemClass C ON A.Item_Number = C.ITEM_NBR
                    and A.Vendor_Number = C.VENDOR_ID
        LEFT JOIN 
			(select MEMBER_NBR, ADDRESSEE, UPDATE_TIMESTAMP, RowNum from ( SELECT MEMBER_NBR, ADDRESSEE, UPDATE_TIMESTAMP,
					ROW_NUMBER() OVER (PARTITION BY MEMBER_NBR ORDER BY UPDATE_TIMESTAMP DESC) AS RowNum
				FROM MBPR01_LOCT_MDB_LOCATION) RankedRecords WHERE RowNum = 1 ) D on A.Member_Number = D.MEMBER_NBR
        JOIN 
        (select VendorNo, ContractID,OSOProgramType, Percentage,Status, ProgramStartDate, CloseDate,EffPeriod,EffClosePeriod
            from (
            select VendorNo, ContractID,OSOProgramType, Percentage, Status, ProgramStartDate, CloseDate,
            CONVERT(VARCHAR(6),ProgramStartDate,112) as EffPeriod,
            CONCAT(YEAR(isnull(CloseDate,DATEADD(Year,1,cast(GETDATE()as date)))),isnull(Format(CloseDate,'MM'),'12')) as EffClosePeriod,
            ROW_NUMBER() OVER (PARTITION BY VENDORNO, OSOProgramType, YEAR(ProgramStartDate) ORDER BY Status ASC, PROGRAMSTARTDATE DESC) AS CONTRACTRANK
            from tblContracts  
            where status in ('Active','Closed') and OSOProgramType is not null and ((Year(ProgramStartDate) <= ? and (Year(CloseDate) >= ? or CloseDate is null)))
                ) AS RankedContracts
        where VendorNo = ?
        and CONTRACTRANK = 1 ) E ON E.VendorNo = A.Vendor_Number and A.ContractID = E.ContractID
        WHERE 
            A.Vendor_Number = ?
            AND A.Year = ?
            AND (
                (E.OSOProgramType = 'NNM' AND A.Project_Type IN ('NW', 'M'))
                OR (E.OSOProgramType = 'GUP' AND A.Type IN ('G'))
                OR (E.OSOProgramType = 'ORIG' AND A.Type IN ('N', 'O'))
                OR (E.OSOProgramType IN ('GUP','ORIG') AND A.Type IN ('G','N','O'))
            )
	) AS Subquery
    WHERE Vendor_Funding IS NOT NULL
	Order by Store_Count, Member_Number, Addressee, Item_Number, Short_Description, Unit_Cost, Eligible_Qty,
        Eligible_Sales, DunsNo, Vendor_Number, Name, Year, PeriodAdded, Type, Vendor_Funding, Credit_Date 
	;
    """
    print("Gathering Vendor Collections")
    collection_query1 = f"""
    WITH VendorCollection AS (
        SELECT
            CASE
                WHEN B.VendorID IS NOT NULL THEN B.VendorID
                ELSE
                    CASE
                        WHEN CHARINDEX('V', B.LongDescription) > 0 THEN
                            SUBSTRING(B.LongDescription, CHARINDEX('V', B.LongDescription) + 1, LEN(B.LongDescription))
                        ELSE
                            NULL
                    END
            END AS VendorID,
            'OSO' AS Program,
            CASE
                WHEN A.DocumentType = 'D' THEN 'Debit'
                ELSE 'Credit'
            END AS PayMethod,
            SUM(A.LineAmount) AS Amount,
            CONVERT(date, A.DocumentDate, 101) AS Date,
            CONCAT(TRIM(A.LineDescription), '_', A.DocumentNbr) AS Reference
        FROM
            tblDocDtl A
        JOIN
            tblDocHdr B
        ON
            A.MbrVndrNbr = B.MbrVndrNbr
            AND A.DocumentNbr = B.DocumentNbr
        WHERE
            A.MbrVndrNbr = ?
            AND A.Account = '130324'
            AND A.Center = '1091159100'
            AND B.ProcessedDate IS NOT NULL
            AND A.DocumentDate > ?
            AND SUBSTRING(A.LineDescription, 1, 4) = ?
        GROUP BY
            A.ApplicationID, A.MbrVndrNbr, A.DocumentNbr, A.DocumentType, A.BatchNbr,
            A.LineNbr, A.LineAmount, A.LineDescription, A.Account, A.Center, A.StmtCd,
            A.UpdateAssocID, A.UpdateDate, A.Approval, A.Account, A.DocumentDate,
            B.VendorID, B.LongDescription
    )
    SELECT *
    FROM VendorCollection
    where VendorID = ?
    ORDER BY VendorID, Date;
        """

    collections_query2 = f"""
    select 
    VID as VendorID,
    'OSO' as Program,
    'Credit' as PayMethod,
    CreditAmt as Amount,
    convert(varchar(10),Date,101) as Date,
    Explanation as Reference
    From MiscCreditMemo
    where VID = ?
    AND Date > ?
    AND DistributionCombo LIKE '%130324%'
    """

    #Use Pandas to execute the queries and read the results into dataframes
    summary = pd.read_sql(summary_query, engine1, params=(Year, Year, Vendor_Number))

    summary_programs = summary[['Programs', '%', 'Program Type', 'Status']]

    details = pd.read_sql(details_query, engine1, params=(Year, Year, Vendor_Number,Vendor_Number, Year))

    #Gather Debits from edoes
    collection_query1 = pd.read_sql(collection_query1, engine2, params=(int(Member_Number), Year, Year, Vendor_Number))

    #Gather Credits from Access OCRDB
    collections_query2 = pd.read_sql(collections_query2, engine3, params=(Vendor_Number, Year))

    #Combing Credits and Debits into one dataframe
    if not collections_query2.empty:
        collections = pd.concat([collection_query1, collections_query2], ignore_index=True)
    else:
        collections = collection_query1



    #Create a currency style
    currency_style = NamedStyle(name='currency',
                                number_format='_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)')

    #Columns I want to format as currency
    currency_details = [6, 8, 15]
    currency_collections = [4]

    #Create the workbook
    workbook = openpyxl.Workbook()

    #Create sheets
    summary_sheet = workbook.create_sheet(title="SUMMARY")
    details_sheet = workbook.create_sheet(title="DETAILS")
    collections_sheet = workbook.create_sheet(title="COLLECTIONS")

    #Remove the default sheet
    del workbook['Sheet']

#Styling SUMMARY SHEET
    # Set background color to white for the entire sheet
    for row in summary_sheet.iter_rows():
        for cell in row:
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    summary_sheet.sheet_view.showGridLines = False

    # Adjust row height & Adjust column width for columns and rows
    summary_sheet.row_dimensions[1].height = 9
    summary_sheet.column_dimensions['A'].width = 1.86
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        summary_sheet.column_dimensions[col_letter].width = 14.86

    # Merge and center cells from column B to J in row 2
    summary_sheet.merge_cells('B2:J2')
    summary_sheet['B2'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['B2'].fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
    summary_sheet['B2'].value = '=DETAILS!L2&" "&"Retail Growth/OSO Detail for"'
    summary_sheet['B2'].font = Font(size=14, bold=True)

    font_size_11 = Font(size=11)

    # Label cell Vendor Infor; Merge and center cells with the formula inside
    summary_sheet['E3'] = 'VENDOR:'
    summary_sheet['E3'].font = Font(bold=True)
    summary_sheet['E3'].alignment = openpyxl.styles.Alignment(horizontal="right")
    summary_sheet.merge_cells('F3:I3')
    summary_sheet['F3'].alignment = openpyxl.styles.Alignment(horizontal="left", indent=5)
    summary_sheet['F3'].value = '=DETAILS!K2'

    summary_sheet['E4'] = 'VID:'
    summary_sheet['E4'].font = Font(bold=True)
    summary_sheet['E4'].alignment = openpyxl.styles.Alignment(horizontal="right")
    summary_sheet.merge_cells('F4:I4')
    summary_sheet['F4'].alignment = openpyxl.styles.Alignment(horizontal="left", indent=5)
    summary_sheet['F4'].value = '=DETAILS!J2'

    # Set background color to RGB(169, 208, 142) for cells from B6 to J6
    for cell in summary_sheet['B6:J6']:
        for c in cell:
            c.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')

    # Labeling Program Summary
    summary_sheet['B6'] = 'Programs'
    summary_sheet['B6'].font = Font(bold=True)
    summary_sheet['B6'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['C6'] = '%'
    summary_sheet['C6'].font = Font(bold=True)
    summary_sheet['C6'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['D6'] = 'Program Type'
    summary_sheet['D6'].font = Font(bold=True)
    summary_sheet['D6'].alignment = openpyxl.styles.Alignment(horizontal="center")


    # Labeling cells G6, H6, I6, and J6
    summary_sheet['G6'] = 'Eligible Sales'
    summary_sheet['G6'].font = Font(bold=True)
    summary_sheet['G6'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['G7'].style = currency_style
    summary_sheet['G7'] = '=SUM(DETAILS!H:H)'
    summary_sheet['H6'] = 'Earned'
    summary_sheet['H6'].font = Font(bold=True)
    summary_sheet['H6'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['H7'].style = currency_style
    summary_sheet['H7'] = '=G7*C7%'
    summary_sheet['I6'] = 'Collections'
    summary_sheet['I6'].font = Font(bold=True)
    summary_sheet['I6'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['I7'].style = currency_style
    summary_sheet['I7'] = '=SUMIF(COLLECTIONS!A:A,$F$4,COLLECTIONS!D:D)'
    summary_sheet['J6'] = 'Balance'
    summary_sheet['J6'].font = Font(bold=True)
    summary_sheet['J6'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['J7'].style = currency_style
    summary_sheet['J7'] = '=H7-I7'


    # Apply borders to the entire dataframe coming out of Summary
#    for row_idx, row in enumerate(summary_sheet.iter_rows(min_row=7, min_col=2), start=7):
#        for cell in row:
#            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
#                                                 right=openpyxl.styles.Side(style='thin'),
#                                                 top=openpyxl.styles.Side(style='thin'),
#                                                 bottom=openpyxl.styles.Side(style='thin'))
            
                    # Apply thin borders to the specified range of cells
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    for col in range(7, 11):
        cell = summary_sheet.cell(row=7, column=col)
        cell.border = thin_border

    #Style the header row with gray background and font size 14
    header_style_details = NamedStyle(name='header_details', font=Font(size=11),
                                      fill=PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid'))
    header_style_collections = NamedStyle(name='header_collections', font=Font(size=11),
                                          fill=PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid'))
#Done Styling Summary Sheet in Excel

    for col_idx in currency_details:
        for row in details_sheet.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = '"$"#,##0.00'

    for col_idx in currency_collections:
        for row in collections_sheet.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = '"$"#,##0.00'

    # For sheet1 (Summary)
    for row_idx, row in enumerate(dataframe_to_rows(summary_programs,index=False, header=False), start=7):
        for col_idx, value in enumerate(row, start=2):
            cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center')
            if 2 <= col_idx <=4:
                cell.border = thin_border


    # For sheet2 (Details) get data and format accordingly
    for row_idx, row in enumerate(dataframe_to_rows(details,index=False, header=False), start=1):
        details_sheet.append(row)

    for col_idx, col_name in enumerate(details.columns, start=1):
        details_sheet.cell(row=1, column=col_idx, value=col_name).style = header_style_details

    # For sheet3 (Collections) get data and format accordingly
    for row_idx, row in enumerate(dataframe_to_rows(collections,index=False, header=False), start=1):
        collections_sheet.append(row)

    for col_idx, col_name in enumerate(collections.columns, start=1):
        collections_sheet.cell(row=1, column=col_idx, value=col_name).style = header_style_collections

    # Get the Vendor Name from the Excel Sheet
    vendor_name_cell = details_sheet['K2']
    if vendor_name_cell.value is not None:
        vendor_name = vendor_name_cell.value.strip()
    else:
        vendor_name = None
    # clean up invalid_chars in Vendor_name so it can save safely
    invalid_chars = '\\/:*?"<>|'
    for char in invalid_chars:
        vendor_name = vendor_name.replace(char, '_')

    # Get the current date
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")

    # Define the directory path
    directory_path = f'H:/MSC/Public/Finance/Reports/Rebates/{Year}/OSO/5. Vendor Detail/'

    # Construct the file name
    file_name = f'{Vendor_Number} {vendor_name} {Year} OSO INV Detail {current_date}.xlsx'

    # Construct the File path based on the year the user selects
    file_path = os.path.join(directory_path, file_name)

    # Save the workbook
    workbook.save(file_path)

    # Open the file
    os.system(f'start excel "{file_path}"')

    print("Successfully Gathering Details and Exporting Excel File now!")

except Exception as e:
    error_message = str(e).encode('unicode_escape').decode('utf-8')
    print(f"An error occurred while opening the file: {error_message}")
except pd.io.sql.DatabaseError as ex:
    print("An error occurred:", ex)
