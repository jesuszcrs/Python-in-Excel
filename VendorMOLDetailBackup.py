import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
import tkinter as tk


script_name = sys.argv[0]
vendor_no = sys.argv[1]
Year = sys.argv[2]

# User provided information
#vendor_no = input("Enter the Vendor Number: ")
#Year = input("Enter Year: ")

# File locations
input_file_path = f'H:/MSC/Public/Finance/Reports/Rebates/{Year}/MOL/Receivable/MOL Receivable {Year}.xlsx'
csv_file_path = f'H:/MSC/Public/Finance/Reports/Rebates/{Year}/MOL/Receivable/MOL Receivable {Year} Detail_Backup_File.csv'
output_file_path = f'H:/MSC/Public/Finance/Reports/Rebates/{Year}/MOL/Vendor Detail/MOL Detail Template.xlsx'
renamed_output_file_path = f'H:/MSC/Public/Finance/Reports/Rebates/{Year}/MOL/Vendor Detail/'
sheet_name_vendor_rec_data = 'New Vendor Rec Data'
sheet_name_member_file = 'New Member File'

# Columns to keep in the output file for 'New Vendor Rec Data'
columns_to_keep_vendor_rec_data = ['MODEL_NBR', 'SHORT_DESC', 'BuyingDept', 'NAME', 'Ven_Cost', 'Period', 'Year', 'Count', 'Contract', 'Contract Percent', 'Receivable']

# Columns to extract from 'New Member File'
columns_to_extract_from_member_file = ['MEMBER_NBR', 'MODEL', 'Period', 'OrderStatus', 'DESC']

# Check if the CSV file already exists
if os.path.exists(csv_file_path):
    # Check modification timestamps of Excel and CSV files
    excel_modified_time = os.path.getmtime(input_file_path)
    csv_modified_time = os.path.getmtime(csv_file_path)

    if excel_modified_time > csv_modified_time:
        # If Excel file is newer than CSV file, read data from Excel and save as CSV
        df_vendor_rec_data = pd.read_excel(input_file_path, sheet_name=sheet_name_vendor_rec_data, header=1)
        df_vendor_rec_data.to_csv(csv_file_path, index=False)
    else:
        # If CSV file is up-to-date, load data from CSV
        df_vendor_rec_data = pd.read_csv(csv_file_path)
else:
    # If the CSV file doesn't exist, read data from Excel and save as CSV
    df_vendor_rec_data = pd.read_excel(input_file_path, sheet_name=sheet_name_vendor_rec_data, header=1)
    df_vendor_rec_data.to_csv(csv_file_path, index=False)

print(f"Connection Successful, begin gathering Vendor Details for V# {vendor_no}")

try:
    # Get user input for vendor number
    vendor_number = vendor_no

    # Check if 'Vendor No' column exists in the dataframe
    if 'Vendor No' in df_vendor_rec_data.columns:
        # Check data types and strip any leading/trailing spaces
        df_vendor_rec_data['Vendor No'] = df_vendor_rec_data['Vendor No'].astype(str).str.strip()
        vendor_number = vendor_number.strip()

        # Filter data based on 'Vendor No' column and Count not equal to 0
        filtered_data_vendor_rec_data = df_vendor_rec_data[(df_vendor_rec_data['Vendor No'] == vendor_number) & (df_vendor_rec_data['Count'] != 0)]

        if filtered_data_vendor_rec_data.empty:
            print("No data found for the provided vendor number.")
        else:
            # Extract unique values from 'MODEL_NBR' column
            model_nbr_list = filtered_data_vendor_rec_data['MODEL_NBR'].unique().tolist()

            # Read data from 'New Member File' sheet
            df_member_file = pd.read_excel(input_file_path, sheet_name=sheet_name_member_file)

            # Filter data from 'New Member File' based on the unique 'MODEL_NBR' list
            filtered_data_member_file = df_member_file[df_member_file['MODEL'].isin(model_nbr_list)]

            # Select only the required columns from 'New Member File'
            filtered_data_member_file = filtered_data_member_file[columns_to_extract_from_member_file]

            # Sort the filtered data by 'MEMBER_NBR' and 'Period'
            filtered_data_member_file = filtered_data_member_file.sort_values(by=['MEMBER_NBR', 'Period'])

            # Write the filtered and sorted data from both sheets to a new sheet in the output Excel file
            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                # Write filtered Vendor Rec Data starting at row 3
                filtered_data_vendor_rec_data[columns_to_keep_vendor_rec_data].to_excel(writer, sheet_name=f'{Year} MOL Detail', index=False, startrow=2)

                print("Gathering Vendor Details")

                # Write filtered Member File data starting below filtered Vendor Rec Data
                start_row_member_file = filtered_data_vendor_rec_data.shape[0] + 5
                filtered_data_member_file.to_excel(writer, sheet_name=f'{Year} MOL Detail', index=False, startrow=start_row_member_file)

                print("Gathering Member Order Details")

                # Access the workbook
                wb = writer.book
                ws = wb[f'{Year} MOL Detail']

                # Define fill color RGB value
                fill_color = "BDD7EE"

                # Create fill pattern
                pattern_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                # Write "Vendor No" to cell A1 and the provided vendor number to cell B1
                ws['A1'] = 'Vendor No'
                ws['B1'] = vendor_number

                # Apply formatting to cell A1 and B1
                for col in range(1, 3):
                    ws.cell(row=1, column=col).fill = pattern_fill
                    ws.cell(row=1, column=col).font = Font(color="000000", bold=True)

                # Apply formatting to header row of filtered_data_vendor_rec_data
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=3, column=col).fill = pattern_fill
                    ws.cell(row=3, column=col).font = Font(color="000000", bold=True)

                # Calculate the starting row for the headers of filtered_data_member_file
                start_row_member_file_headers = filtered_data_vendor_rec_data.shape[0] + 6

                # Apply formatting to headers of filtered_data_member_file
                for col in range(1, filtered_data_member_file.shape[1] + 1):
                    ws.cell(row=start_row_member_file_headers, column=col).fill = pattern_fill
                    ws.cell(row=start_row_member_file_headers, column=col).font = Font(color="000000", bold=True)
                
                # Calculate the maximum length of data in each column
                max_length = {}
                for column in ws.columns:
                    max_length[column[0].column_letter] = max(len(str(cell.value)) for cell in column)

                # Set the column width based on the maximum length of data in each column
                for column, length in max_length.items():
                    ws.column_dimensions[column].width = length

            # clean up invalid_chars in Vendor_name so it can save safely
            vendor_name = ws['D4'].value.strip()
            invalid_chars = '\\/:*?"<>|'
            for char in invalid_chars:
                vendor_name = vendor_name.replace(char, '_')

            # Create Vendor Detail file name
            current_date = datetime.now().strftime('%Y-%m-%d')
            vendor_detail_file = f'V{vendor_no} {vendor_name} {Year} MOL Detail {current_date}.xlsx'
            detail_file_path = os.path.join(renamed_output_file_path,vendor_detail_file)

            # Remove existing file if exists
            if os.path.exists(detail_file_path):
                os.remove(detail_file_path)

            os.rename(output_file_path,detail_file_path)

            print(f"Filtered data written to '{detail_file_path}'")

            print("Successfully Gathered Details and Exporting Excel File Now!")

            # Open the new file
            os.system(f'start excel "{detail_file_path}"')
    else:
        print("No 'Vendor No' column found in the dataframe.")
except Exception as e:
    print(f"An error occurred: {e}")
