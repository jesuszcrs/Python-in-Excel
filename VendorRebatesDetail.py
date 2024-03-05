import datetime
import os
import subprocess
from urllib.parse import quote_plus

import pandas as pd
import numpy as np
import pyodbc
import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import create_engine
import sys

script_name = sys.argv[0]
Vendor_Number = sys.argv[1]
Year = sys.argv[2]

#Vendor_Number = input("Enter Vendor Number: ")
#Year = input("Enter Year: ")

params = {
    'Vendor_no': Vendor_Number,
    'Year': Year,
    'Member_Number': None,
    'ProgramType': [],
    'Account_Number': []
}

#Define your first SQL Server Connection Parameters
#Credentials removed for privacy

#Create a connection string for the first SQL Server
connection_string1 = f"mssql+pyodbc://{username1}:{password1}@{server1}/{database1}?driver=SQL+Server"

#Define your second SQL Server Connection Parameters
#Credentials removed for privacy

#Create a connection string for the second SQL Server
connection_string2 = f"mssql+pyodbc://@{server2}/{database2}?driver=SQL+Server"

#Define your third SQL Server Connection string for connection3 (Access database)
#Credentials removed for privacy

connection_string3 = f"mssql+pyodbc://{UID}:{PWD}@{DSN}"
print (f"Connection Successful, begin gathering Vendor Details for V# {Vendor_Number}")
try:
    # Create a SQLAlchemy engine for the first database connection
    engine1 = create_engine(connection_string1)

    #Create a SQLAlchemy engine for the second database connection
    engine2 = create_engine(connection_string2)

    #Establish a connection to the third Access database
    engine3 = create_engine(connection_string3)

    # Execute SQL query to pull account nbr in eDOES for odd or even year
    if int(Year) % 2 == 0:  # Even Year
        accountnbr_query = "SELECT * FROM datAccountCodes where ProgramType IN ('ADV','FUNC','RIF','INCENT','RMKTF') and yearmod = 2"
    else:  # ODD Year
        accountnbr_query = "SELECT * FROM datAccountCodes where ProgramType IN ('ADV','FUNC','RIF','INCENT','RMKTF') and yearmod = 1"

    account_nbr_data = pd.read_sql(accountnbr_query, engine1)
    if not account_nbr_data.empty:
        Account_Number = [str(account) for account in account_nbr_data['AccountNbr']]
        ProgramType = [str(program_type)[:20] for program_type in account_nbr_data['ProgramType']]
        params['Account_Number'] = Account_Number
        params['ProgramType'] = ProgramType
    else:
        print("No results found for AccountNbr and ProgramType.")


    # Execute SQL query to pull APDunsNo for Vendor from tblVendorPR table
    apdunsno_query = "Select DunsNo from tblVendorPR where VendorNo = ?"

    apdunsno = pd.read_sql(apdunsno_query, engine1, params=(Vendor_Number,))

    if not apdunsno.empty:
        Member_Number = apdunsno.iloc[0]['DunsNo']
        params['Member_Number'] = Member_Number
    else:
        print("Member_number not found in Vendor_Pr for the given Vendor Number.")

    summary_query1 = f""" select 
    concat(ProgramType,' ', ?) as ProgramType,
    isnull(Percentage/100,0) as Percentage,
    Type, 
    isnull(FlatAmount,0) as FlatAmount
    from tblRebateReportsByPeriod 
    where ProgramType in ('ADV','FUNC','RIF','RMKTF') 
    AND EffPeriod = (select max(effperiod) from tblRebateReportsByPeriod where VendorNo = ? and Year = ?)
    AND Vendorno = ? and Year = ?
    order by ProgramType
    """

    summary_query2 = f"""
    Select A.Incentive/100 as Percentage,A.PurchaseType,
    case when A.BoundType = 'A' then A.LowerBound when A.BoundType = 'AYOY' then A.TOT_PURCHASES_PRIOR + A.LowerBound 
        else ((A.LowerBound/100) + 1) * A.TOT_PURCHASES_PRIOR end as LowerBoundAmt, 
    iif(A.UpperBound is null, 99999999.99, case when A.BoundType = 'A' then A.UpperBound - .01 
            when A.BoundType = 'AYOY' then A.TOT_PURCHASES_PRIOR + A.UpperBound
            else (((A.UpperBound/100) + 1) * A.TOT_PURCHASES_PRIOR) - .01 end) as UpperBoundAmt
    From DW_Work.dbo.tblRebatesByYear A
            JOIN 
            (select VendorNo, ContractID,ProgramType, Percentage,Status, ProgramStartDate, CloseDate,EffPeriod,EffClosePeriod
                from (
                select VendorNo, ContractID,ProgramType, Percentage, Status, ProgramStartDate, CloseDate,
                CONVERT(VARCHAR(6),ProgramStartDate,112) as EffPeriod,
                CONCAT(YEAR(isnull(CloseDate,DATEADD(Year,1,cast(GETDATE()as date)))),isnull(Format(CloseDate,'MM'),'12')) as EffClosePeriod,
                ROW_NUMBER() OVER (PARTITION BY VENDORNO, YEAR(ProgramStartDate) ORDER BY Status ASC, PROGRAMSTARTDATE DESC) AS CONTRACTRANK
                from tblContracts  
                where status in ('Active','Closed') and ProgramType = 'INCENT' and ((Year(ProgramStartDate) <= ? and (Year(CloseDate) >= ? or CloseDate is null)))
                    ) AS RankedContracts
            where VendorNo = ?
            and CONTRACTRANK = 1 ) B ON B.VendorNo = A.VendorNo and A.ContractID = B.ContractID
    Where (A.INVOICE_YEAR = ?  or A.INVOICE_YEAR is null) and A.ProcessStatus = 'Include' 
    Order By A.ProgramTierID, A.TierNumber Asc 
    """

    retroactive_query = f"""
    select case
    when RetroactiveTiers = 1 THEN 'Y'
    else 'N'
    end as Retroactive
    from tblRebatesByYear
    where VendorNo = ? and Invoice_Year = ?
    and ProcessStatus = 'include'
    group by RetroactiveTiers
    """
    print("Gathering Vendor Details")
    details_query1 = f"""
    select B.DunsNo AS AP_VENDOR_NBR,
    A.VENDOR_ID AS MSC_VENDOR_ID, 
    B.NAME AS VENDOR_NAME,
    b.BuyingDept AS DEPARTMENT,
    A.ITEM_NBR AS TV_ITEM_NBR,
    COALESCE(C.SHORT_DESCRIPTION, ' ') AS ITEM_DESCRIPTON,
    CONVERT(VARCHAR(10), A.DATE_RECEIVED, 101) as INV_DATE,
    CONVERT(VARCHAR(10), A.DATE_RECEIVED, 101) AS EFF_DATE,
    CONVERT(VARCHAR(10), A.PO_DATE, 101) AS PO_DATE,
    ' ' AS INVOICE_NBR,
    A.PO_NBR,
    A.FISCAL_PERIOD AS INV_PERIOD,
    A.FISCAL_YEAR AS YEAR,
    SUM(A.STK-A.STK) AS DS, 
    sum(STK) AS STK,
    sum(STK) AS TOTAL
    
    FROM DW_Work.DBO.IMPORTS A INNER JOIN DW_Work.DBO.tblVendorPR B ON A.VENDOR_ID = B.VendorNo
    LEFT JOIN DW_Work.DBO.tblItemClass C ON A.ITEM_NBR = C.ITEM_NBR
    
    WHERE a.vendor_id = ?
    AND a.FISCAL_YEAR = ?
    
    GROUP BY B.DunsNo, A.VENDOR_ID, A.FISCAL_PERIOD, A.ITEM_NBR, b.NAME,
    a.DATE_RECEIVED, a.FISCAL_YEAR, A.STK, b.BuyingDept, C.SHORT_DESCRIPTION,
    A.PO_DATE, A.PO_NBR
         """

    details_query2 = f"""
    select A.AP_VENDOR_NBR,
    A.VENDOR_ID AS MSC_VENDOR_ID, 
    B.NAME AS VENDOR_NAME, 
    b.BuyingDept AS DEPARTMENT,
    A.ITEM_NBR AS TV_ITEM_NBR,
    COALESCE(C.SHORT_DESCRIPTION, ' ') AS ITEM_DESCRIPTON, 
    CONVERT(varchar(10), A.INV_DATE, 101) AS INV_DATE,
    CONVERT(varchar(10), A.EFF_DATE,101) AS EFF_DATE,
    ' ' AS PO_DATE,
    A.INVOICE_NBR,
    ' ' AS PO_NBR,
    A.INV_PERIOD,
    A.INVOICE_YEAR AS YEAR,
    SUM(A.DS_AMT+DS_SPLIT_AMT) AS DS,
    sum(STK_AMT) AS STK,
    SUM(A.DS_AMT+DS_SPLIT_AMT+STK_AMT) AS TOTAL
    
    FROM DW_Work.DBO.VRDT_VDR_REB_DTL A INNER JOIN DW_Work.DBO.tblVendorPR B ON A.VENDOR_ID = B.VendorNo 
    LEFT JOIN DW_Work.DBO.tblItemClass C ON A.ITEM_NBR = C.ITEM_NBR
    
    WHERE a.vendor_id = ?
    AND a.invoice_year = ?
    
    GROUP BY A.AP_VENDOR_NBR, A.VENDOR_ID, A.INVOICE_NBR, A.INV_PERIOD,
    A.ITEM_NBR, b.NAME, a.INV_DATE, a.eff_date, a.INVOICE_YEAR,
    b.BuyingDept, C.SHORT_DESCRIPTION, A.ITEM_DESCRIPTION
         """
    
    print("Gathering Vendor Collections")    
    collections_query1 = f"""
    select MSC_VENDOR_ID, CONCAT(PROGRAMTYPE,' ',REBATE_YEAR) as Program,
    case when DOCUMENT_TYPE = 'DM' then 'Debit'
    when DOCUMENT_TYPE = 'CM' then 'Credit'
    when DOCUMENT_TYPE = 'PB' then 'Payback'
    when DOCUMENT_TYPE = 'RC' then 'Reclass'
    when DOCUMENT_TYPE = 'CK' then 'Check'
    when DOCUMENT_TYPE = 'DR' then 'Debit'
    end as PayMethod, 
    format(DOCUMENT_DATE, 'yyyy-MM-dd') as Date,
    DOCUMENT_AMOUNT,
    CONCAT(COMMENT,' INV# ',DOCUMENT_NBR) as Reference
    from datCollections
    where MSC_VENDOR_ID = ? AND REBATE_YEAR = ? and trim(PROGRAMTYPE) IN ({', '.join(['?'] * len(ProgramType))})
    order by MSC_VENDOR_ID, PROGRAMTYPE,DOCUMENT_DATE
    """

    collections_query2 = f"""
    select 
    VID as VendorID,
        CASE 
        WHEN CHARINDEX('FUNC', Explanation) > 0 THEN 'FUNC'
        WHEN CHARINDEX('ADV', Explanation) > 0 THEN 'ADV'
        WHEN CHARINDEX('RIF', Explanation) > 0 THEN 'RIF'
        WHEN CHARINDEX('INC', Explanation) > 0 THEN 'INCENT'
        WHEN CHARINDEX('RMKTF', Explanation) > 0 THEN 'RMKTF'
        ELSE ''
        End as Program,
    'Credit' as PayMethod,
    convert(varchar(10),Date,101) as Date,
    -CreditAmt as Amount,
    Explanation as Reference
    From MiscCreditMemo
    where VID = ?
    AND Date > ?
    """


    #Use Pandas to execute the queries and read the results into dataframes
    summary = pd.read_sql(summary_query1, engine1, params=(Year, Vendor_Number, Year, Vendor_Number, Year))

    incentive = pd.read_sql(summary_query2, engine1, params=(Year, Year, Vendor_Number, Year))

    retroactive = pd.read_sql(retroactive_query, engine1, params=(Vendor_Number, Year))

    #Gather Import Details from DW_Work
    details_query1 = pd.read_sql(details_query1, engine1, params=(Vendor_Number, Year))

    #Gather Domestic Details form DW_Work
    details_query2 = pd.read_sql(details_query2, engine1, params=(Vendor_Number, Year))

    #Combine Import & Domestic Details from DW_Work
    if not details_query1.empty:
        details = pd.concat([details_query1, details_query2], ignore_index=True)
    else:
        details = details_query2

    #Gather Debits from edoes
    collections_query1 = pd.read_sql(collections_query1, engine1, params=(Vendor_Number, Year, *ProgramType))

    #Gather Credits from Access OCRDB
    collections_query2 = pd.read_sql(collections_query2, engine3, params=(Vendor_Number, Year))

    #Combing Credits and Debits into one dataframe
    if not collections_query2.empty:
        collections = pd.concat([collections_query1, collections_query2], ignore_index=True)
    else:
        collections = collections_query1

    #Create a currency style
    currency_style = NamedStyle(name='currency',
                                number_format='"$"#,##0.00_);("$"#,##0.00;("$"#,##0.00;"-"??_)')
    
    #Columns I want to format as currency
    currency_details = [13, 14, 15]
    currency_collections = [5]

    #Create the workbook
    workbook = openpyxl.Workbook()

    #Create sheets
    summary_sheet = workbook.create_sheet(title="SUMMARY")
    details_sheet = workbook.create_sheet(title="DETAILS")
    collections_sheet = workbook.create_sheet(title="COLLECTIONS")

    #Remove the default sheet
    del workbook['Sheet']

#Styling SUMMARY SHEET
    # Set background color to white for the entire sheet
    for row in summary_sheet.iter_rows():
        for cell in row:
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    summary_sheet.sheet_view.showGridLines = False

    # Adjust row height & Adjust column width for columns and rows
    summary_sheet.row_dimensions[1].height = 9
    summary_sheet.column_dimensions['A'].width = 1.86
    summary_sheet.column_dimensions['H'].width = 1.86
    summary_sheet.column_dimensions['B'].width = 4.57
    summary_sheet.column_dimensions['C'].width = 4.57
    for col_letter in ['D', 'E', 'F', 'G', 'I', 'J', 'K','L']:
        summary_sheet.column_dimensions[col_letter].width = 14.86

    # Merge and center cells from column B to J in row 2
    summary_sheet.merge_cells('B2:L2')
    summary_sheet['B2'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['B2'].fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
    summary_sheet['B2'].value = '=DETAILS!M2&" "&"Rebate Invoice Detail for"'
    summary_sheet['B2'].font = Font(size=14, bold=True)

    font_size_11 = Font(size=11)

     # Label cell Vendor Infor; Merge and center cells with the formula inside
    summary_sheet['F3'] = 'VENDOR:'
    summary_sheet['F3'].font = Font(bold=True)
    summary_sheet['F3'].alignment = openpyxl.styles.Alignment(horizontal="right")
    summary_sheet.merge_cells('G3:K3')
    summary_sheet['G3'].alignment = openpyxl.styles.Alignment(horizontal="left", indent=5)
    summary_sheet['G3'].value = '=DETAILS!C2'

    summary_sheet['F4'] = 'VID:'
    summary_sheet['F4'].font = Font(bold=True)
    summary_sheet['F4'].alignment = openpyxl.styles.Alignment(horizontal="right")
    summary_sheet.merge_cells('G4:K4')
    summary_sheet['G4'].alignment = openpyxl.styles.Alignment(horizontal="left", indent=5)
    summary_sheet['G4'].value = '=DETAILS!B2'

    # Set background color to RGB(169, 208, 142) for cells from B6 to J6
    for cells_range in ['B6:L6', 'D13:L13']:
        for cell in summary_sheet[cells_range]:
            for c in cell:
                c.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
    
    # Labeling Program Summary
    summary_sheet.merge_cells('B6:C6')
    summary_sheet['B6'] = 'Programs'
    summary_sheet['B6'].font = Font(bold=True)
    summary_sheet['B6'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['D6'] = '%'
    summary_sheet['D6'].font = Font(bold=True)
    summary_sheet['D6'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['E6'] = 'Purch Type'
    summary_sheet['E6'].font = Font(bold=True)
    summary_sheet['E6'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['F6'] = 'Flat $'
    summary_sheet['F6'].font = Font(bold=True)
    summary_sheet['F6'].alignment = openpyxl.styles.Alignment(horizontal="center")

    # Define the range of cells where you want to apply the formulas for the Rebate Calculations
    rebate_start_row = 7
    incentive_start_row = 14
    rebate_end_row = 11
    incentive_end_row = 19
    columns_with_formula = {
        'I': '=IF(E{}="DSO",SUM(DETAILS!N:N),IF(E{}="WHS",SUM(DETAILS!O:O),IF(E{}="TOT",SUM(DETAILS!P:P),0)))',
        'J': '=IF(F{}>0.01,F{}, IFERROR(I{}*D{},0))',
        'K': '=SUMIF(COLLECTIONS!B:B,$B{},COLLECTIONS!E:E)',
        'L': '=J{}-K{}'
    }

    #Apply formulas to the Rebate Range over each column and apply the formula to the specified range of cells
    for col, formula in columns_with_formula.items():
        for row in range(rebate_start_row, rebate_end_row + 1):
            cell = summary_sheet[col + str(row)]
            # Apply the formula to the cell without modifying it
            cell.value = formula.format(row, row, row, row)
            cell.style = 'Currency'  # Make sure to define 'Currency' style
            cell.alignment = Alignment(horizontal="center")
    
    #Apply formulas to the Incentive Range over each column and apply the formula to the specified range of cells
    for col, formula in columns_with_formula.items():
        for row in range(incentive_start_row, incentive_end_row + 1):
            cell = summary_sheet[col + str(row)]
            # Apply the formula to the cell without modifying it
            cell.value = formula.format(row, row, row, row)
            cell.style = 'Currency'
            cell.alignment = Alignment(horizontal="center")

    # Labeling cells G6, H6, I6, and J6
    labels = ['Purchases', 'Earned', 'Collections', 'Balance']
    columns = ['I', 'J', 'K', 'L']
    rows = [6, 13]
    for row in rows:
        for col, label in zip(columns, labels):
            cell = summary_sheet[col + str(row)]
            cell.value = label
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Apply thin borders to the specified range of cells
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
    # Apply border to the specified range of cells
    for col in columns_with_formula.keys():
        for row in range(rebate_start_row, rebate_end_row + 1):
            cell = summary_sheet[col + str(row)]
            cell.border = thin_border
    for col in columns_with_formula.keys():
        for row in range(incentive_start_row, incentive_end_row + 1):
            cell = summary_sheet[col + str(row)]
            cell.border = thin_border

    #Style the header row with gray background and font size 14
    header_style_details = NamedStyle(name='header_details', font=Font(size=11),
                                      fill=PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid'))
    header_style_collections = NamedStyle(name='header_collections', font=Font(size=11),
                                          fill=PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid'))

#Done Styling Summary Sheet in Excel

    for col_idx in currency_details:
        for row in details_sheet.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = '"$"#,##0.00'

    for col_idx in currency_collections:
        for row in collections_sheet.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = '"$"#,##0.00'

    # Define the range of rows and columns where you want to populate the data
    summary1_start_row = 7
    summary1_end_row = 11
    summary1_columns = ['B', 'C', 'D', 'E', 'F']
    
    # Predefined list of programs with dynamic year
    programs = [f'ADV {Year}', f'FUNC {Year}', f'FUNC2 {Year}', f'RIF {Year}', f'RMKTF {Year}']

    # Populate the predefined programs in column B
    for idx, program in enumerate(programs, start=summary1_start_row):
       #Merge cells B7:C7 from rows 7 to 11 for ProgramTypes
        summary_sheet.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=3)
        #Set the value for the leftmost cell of the merged range
        summary_sheet[f'B{idx}'].value = program

    # Iterate over the data and populate the cells
    for idx, row in summary.iterrows():
        program = row['ProgramType']
        percentage = row['Percentage']
        type_value = row['Type']
        flat_amount = row['FlatAmount']

        # Determine the row index to populate based on the program
        row_idx = programs.index(program) + summary1_start_row

        # Populate the cells in columns C to E
        summary_sheet[f'D{row_idx}'].value = percentage
        summary_sheet[f'E{row_idx}'].value = type_value
        summary_sheet[f'F{row_idx}'].value = flat_amount

    # Apply formatting and border styles
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    for col in summary1_columns:
        for idx in range(summary1_start_row, summary1_end_row + 1):
            cell = summary_sheet[f'{col}{idx}']
            if col == 'B':  # Left alignment for ProgramType
                cell.alignment = Alignment(horizontal='center')
                cell.font = openpyxl.styles.Font(bold=True)
            elif col == 'D':  # Center alignment and percent format for Percentage
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = '0.00%'
            else:  # Center alignment for Purch Type and Flat Amount
                cell.alignment = Alignment(horizontal='center')
            
            # Apply thin border style to specific cells
            cell.border = thin_border

#Styliing Incentive Summary
    # Labeling Incentive Program Summary
    summary_sheet.merge_cells('B14:C14')
    summary_sheet['D13'] = '%'
    summary_sheet['D13'].font = Font(bold=True)
    summary_sheet['D13'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['E13'] = 'Purch Type'
    summary_sheet['E13'].font = Font(bold=True)
    summary_sheet['E13'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['F13'] = 'Tier Start'
    summary_sheet['F13'].font = Font(bold=True)
    summary_sheet['F13'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['G13'] = 'Tier End'
    summary_sheet['G13'].font = Font(bold=True)
    summary_sheet['G13'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['B14'] = f"Incentive {Year}"
    summary_sheet['B14'].font = Font(bold=True)
    summary_sheet['B14'].alignment = openpyxl.styles.Alignment(horizontal="center")
    summary_sheet['B15'] = "$1"
    summary_sheet['B15'].font = Font(bold=True)
    summary_sheet['B15'].alignment = openpyxl.styles.Alignment(horizontal="center")  


    # For the retroactive flag define specific cell to populate flag
    specific_row = 15
    specific_column = 3

    if not retroactive.empty:
        retroactive_value = retroactive["Retroactive"].iloc[0]
    else:
        retroactive_value = None

    cell = summary_sheet.cell(row=specific_row, column=specific_column)
    cell.value = retroactive_value
    #Apply thin borders to the range B14 to C15
    for row in range(14, 16):
        for col in range(2, 4):
            cell = summary_sheet.cell(row=row, column=col)
            cell.border = thin_border

    # Fill in empty rows in Incentive with formatting for when there's no Incentive data
    for row_idx in range(14, 20):  # Rows 14 to 19
        for col_idx in range(4, 8):  # Columns D to G
            cell = summary_sheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx == 4:  # Checking the column index
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = '0.00%'  # Apply percentage format
            elif col_idx == 5:  # Checking the column index
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='center')
                cell.style = 'Currency'  # Apply currency format

    # For sheet1 (Incentives in Summary)
    for row_idx, row in enumerate(dataframe_to_rows(incentive, index=False, header=False), start=14):
        for col_idx, value in enumerate(row, start=4):
            cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 4:  # Checking the column index
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = '0.00%'  # Apply percentage format
            elif col_idx == 5:  # Checking the column index
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='center')
                cell.style = 'Currency'  # Apply currency format
            cell.border = thin_border
    # Reapply thin borders to columns F and G after populating the cells with data
    for row_idx in range(14, 20):  # Rows 14 to 19
        for col_idx in range(4, 8):  # Columns D to G
            cell = summary_sheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # For sheet2 (Details) get data and format accordingly
    for row_idx, row in enumerate(dataframe_to_rows(details,index=False, header=False), start=1):
        details_sheet.append(row)

    for col_idx, col_name in enumerate(details.columns, start=1):
        details_sheet.cell(row=1, column=col_idx, value=col_name).style = header_style_details

    # For sheet3 (Collections) get data and format accordingly
    for row_idx, row in enumerate(dataframe_to_rows(collections,index=False, header=False), start=1):
        collections_sheet.append(row)

    for col_idx, col_name in enumerate(collections.columns, start=1):
        collections_sheet.cell(row=1, column=col_idx, value=col_name).style = header_style_collections

    # Get the Vendor Name from the Excel Sheet
    vendor_name = details_sheet['C2'].value.strip()

    # clean up invalid_chars in Vendor_name so it can save safely
    invalid_chars = '\\/:*?"<>|'
    for char in invalid_chars:
        vendor_name = vendor_name.replace(char, '_')

    # Get the current date
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")

    # Define the directory path
    directory_path = f'H:/MSC/Public/Finance/Reports/Rebates/{Year}/Rebate Earnings/Invoice Detail/'

    # Construct the file name
    file_name = f'{Vendor_Number} {vendor_name} {Year} INV Detail {current_date}.xlsx'

    # Construct the File path based on the year the user selects
    file_path = os.path.join(directory_path, file_name)

    # Save the workbook
    workbook.save(file_path)

    # Open the file
    os.system(f'start excel "{file_path}"')

    print("Successfully Gathering Details and Exporting Excel File now!")

except Exception as e:
    error_message = str(e).encode('unicode_escape').decode('utf-8')
    print(f"An error occurred while opening the file: {error_message}")
except pd.io.sql.DatabaseError as ex:
    print("An error occurred:", ex)
